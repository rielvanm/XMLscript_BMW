import sys
import time
import traceback
import threading
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import tkinter as tk
from tkinter import filedialog, messagebox, ttk


SCRIPT_VERSION = "V3.0"
DESIGNER       = "MvR"
DEPARTMENT     = "tdOob"


def resource_path(relative_path: str) -> Path:
    base_path = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base_path / relative_path
def strip_namespaces(root: ET.Element) -> None:
    for el in root.iter():
        if "}" in el.tag:
            el.tag = el.tag.split("}", 1)[1]


def safe_float(x: str):
    try:
        return float(x)
    except Exception:
        return None


def update_min_max(cur_min, cur_max, v):
    if v is None:
        return cur_min, cur_max
    if cur_min is None or v < cur_min:
        cur_min = v
    if cur_max is None or v > cur_max:
        cur_max = v
    return cur_min, cur_max


def parse_meta_xml(meta_path: str) -> dict:
    meta_file = Path(meta_path)
    if not meta_file.exists():
        return {}

    tree = ET.parse(meta_file)
    root = tree.getroot()
    strip_namespaces(root)

    return {
        "VIN17": (root.findtext("VIN17") or "").strip(),
        "Baureihe": (root.findtext("BasicFeatures/Baureihe") or "").strip(),
        "Ereihe": (root.findtext("BasicFeatures/Ereihe") or "").strip(),
        "VerkaufsBezeichnung": (root.findtext("BasicFeatures/VerkaufsBezeichnung") or "").strip(),
        "Motor": (root.findtext("BasicFeatures/Motor") or "").strip(),
        "Getriebe": (root.findtext("BasicFeatures/Getriebe") or "").strip(),
        "Modelljahr": (root.findtext("BasicFeatures/Modelljahr") or "").strip(),
        "Marke": (root.findtext("BasicFeatures/Marke") or "").strip(),
        "TypeCode": (root.findtext("BasicFeatures/TypeCode") or "").strip(),
        "StartDate": (root.findtext("StartDate") or "").strip(),
        "EndDate": (root.findtext("EndDate") or "").strip(),
        "WorkState": (root.findtext("WorkState") or "").strip(),
        "VehicleCommunication": (root.findtext("VehicleCommunication") or "").strip(),
        "ComputerName": (root.findtext("ComputerName") or "").strip(),
        "UserName": (root.findtext("UserName") or "").strip(),
        "IstaCaseId": (root.findtext("IstaCaseId") or "").strip(),
        "DateOfFastaRead": (root.findtext("DateOfFastaRead") or "").strip(),
        "DistanceOfFastaRead": (root.findtext("DistanceOfFastaRead") or "").strip(),
        "DealerNumber": (root.findtext("DealerNumber") or "").strip(),
        "distributionPartnerNumber": (root.findtext("distributionPartnerNumber") or "").strip(),
        "outletNumber": (root.findtext("outletNumber") or "").strip(),
    }


def extract_vin(xml_path: str) -> str:
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        strip_namespaces(root)
        return (root.findtext("VCI/VIN") or root.findtext("VIN17") or "").strip()
    except Exception:
        return ""


def pause_exit(msg: str, seconds: int = 6) -> None:
    print(msg)
    print(f"\nVenster sluit automatisch over {seconds} seconden...")
    time.sleep(seconds)


def write_log(out_dir: Path, text: str) -> None:
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "BMW_XML_debug.log").write_text(text, encoding="utf-8", errors="replace")
    except Exception:
        pass


def pick_paths_from_argv(argv: list[str]) -> tuple[str, str]:
    meta = ""
    trans = ""
    for a in argv:
        p = Path(a)
        name = p.name.upper()
        if "RG_META" in name and p.suffix.lower() == ".xml":
            meta = str(p)
        if "RG_TRANS" in name and p.suffix.lower() == ".xml":
            trans = str(p)
    return meta, trans


def parse_vehicle_xml_to_excel(
    trans_xml_path: str,
    meta_xml_path: str,
    out_excel_path: str,
    sort_mode: str = "time",  # time/km
    f_ort_width: int = 80,
    wrap_f_ort_text: bool = True,
) -> None:
    trans_file = Path(trans_xml_path)
    if not trans_file.exists():
        raise FileNotFoundError(f"RG_TRANS niet gevonden: {trans_file}")

    tree = ET.parse(trans_file)
    root = tree.getroot()
    strip_namespaces(root)

    vin = (root.findtext("VCI/VIN") or root.findtext("VIN17") or "").strip()
    brand = (root.findtext("BrandName") or "").strip()

    meta = parse_meta_xml(meta_xml_path) if meta_xml_path else {}

    ecu_rows: list[dict] = []
    dtc_rows: list[dict] = []
    dtc_ctx_rows: list[dict] = []
    zfs_result_rows: list[dict] = []

    for ecu in root.findall("ECU/ECU"):
        ec_title = ecu.findtext("ECUTitle", default="")
        variante = ecu.findtext("VARIANTE", default="")
        bus = ecu.findtext("BUS", default="")
        ecu_grobname = ecu.findtext("ECU_GROBNAME", default="")
        ecu_name = ecu.findtext("ECU_NAME", default="")
        ecu_sgbd = ecu.findtext("ECU_SGBD", default="")
        ecu_gruppe = ecu.findtext("ECU_GRUPPE", default="")
        ecu_adr = ecu.findtext("ECU_ADR", default="")
        diag_proto = ecu.findtext("DiagProtocoll", default="")

        comm_ok = ecu.findtext("COMMUNICATION_SUCCESSFULLY", default="")
        ident_ok = ecu.findtext("IDENT_SUCCESSFULLY", default="")
        fs_ok = ecu.findtext("FS_SUCCESSFULLY", default="")
        is_ok = ecu.findtext("IS_SUCCESSFULLY", default="")
        serial_ok = ecu.findtext("SERIAL_SUCCESSFULLY", default="")
        svk_ok = ecu.findtext("SVK_SUCCESSFULLY", default="")

        id_lief_nr = ecu.findtext("ID_LIEF_NR", default="")
        id_lief_text = ecu.findtext("ID_LIEF_TEXT", default="")
        seriennummer = ecu.findtext("SERIENNUMMER", default="")

        prog_datum, prog_km = "", ""
        svk = ecu.find("SVK")
        if svk is not None:
            prog_datum = svk.findtext("PROG_DATUM", default="")
            prog_km = svk.findtext("PROG_KM", default="")

        ecu_rows.append(
            {
                "VIN": vin,
                "BrandName": brand,
                "ECUTitle": ec_title,
                "VARIANTE": variante,
                "ECU_GROBNAME": ecu_grobname,
                "ECU_NAME": ecu_name,
                "ECU_SGBD": ecu_sgbd,
                "ECU_GRUPPE": ecu_gruppe,
                "BUS": bus,
                "ECU_ADR": ecu_adr,
                "DiagProtocoll": diag_proto,
                "ID_LIEF_NR": id_lief_nr,
                "ID_LIEF_TEXT": id_lief_text,
                "SERIENNUMMER": seriennummer,
                "PROG_DATUM": prog_datum,
                "PROG_KM": prog_km,
                "COMMUNICATION_SUCCESSFULLY": comm_ok,
                "IDENT_SUCCESSFULLY": ident_ok,
                "FS_SUCCESSFULLY": fs_ok,
                "IS_SUCCESSFULLY": is_ok,
                "SERIAL_SUCCESSFULLY": serial_ok,
                "SVK_SUCCESSFULLY": svk_ok,
            }
        )

        def add_dtcs(container_tag: str):
            container = ecu.find(container_tag)
            if container is None:
                return

            for dtc in container.findall("DTC"):
                ecu_addr = dtc.findtext("ecuAddress", default="")
                dtc_id = dtc.findtext("dtcId", default="")
                f_ort = dtc.findtext("F_ORT", default="")
                f_ort_text = dtc.findtext("F_ORT_TEXT", default="")
                f_art = dtc.findtext("F_ART", default="")
                f_vorhanden_text = dtc.findtext("F_VORHANDEN_TEXT", default="")
                f_warnung_text = dtc.findtext("F_WARNUNG_TEXT", default="")
                relevance = dtc.findtext("Relevance", default="")
                id_field = dtc.findtext("Id", default="")

                dtc_context = dtc.find("DTCContext")

                # --- CASE 1: Geen context -> wél 1 DTC-regel wegschrijven ---
                if dtc_context is None:
                    dtc_rows.append(
                        {
                            "F_UW_ZEIT": "",
                            "F_UW_KM": "",
                            "F_ORT_TEXT": f_ort_text,
                            "Speed_Min_kmh": None,
                            "Speed_Max_kmh": None,
                            "RPM_Min": None,
                            "RPM_Max": None,
                            "BatteryV_Min": None,
                            "BatteryV_Max": None,
                            "GPS_Time_UTC": None,
                            "CustomerTime": None,
                            "TimeSinceStart_s": None,
                            "Throttle_pct": None,
                            "EngineTemp_C": None,
                            "IntakeTemp_C": None,
                            "Occurrence": None,
                            "F_UW_ANZ": "",
                            "VIN": vin,
                            "BrandName": brand,
                            "ECUTitle": ec_title,
                            "ECU_GROBNAME": ecu_grobname,
                            "ECU_NAME": ecu_name,
                            "VARIANTE": variante,
                            "ECU_ADR": ecu_adr,
                            "Container": container_tag,
                            "ecuAddress": ecu_addr,
                            "dtcId": dtc_id,
                            "Id": id_field,
                            "Relevance": relevance,
                            "F_ORT": f_ort,
                            "F_ART": f_art,
                            "F_VORHANDEN_TEXT": f_vorhanden_text,
                            "F_WARNUNG_TEXT": f_warnung_text,
                        }
                    )
                    # BELANGRIJK: niet 'return' (anders stop je alle volgende DTC's in dit containerblok)
                    continue

                # Meerdere occurrences per DTC
                for occ_idx, type_ctx in enumerate(dtc_context.findall("typeDTCContext"), start=1):
                    ctx_km = type_ctx.findtext("F_UW_KM", default="")
                    ctx_zeit = type_ctx.findtext("F_UW_ZEIT", default="")
                    ctx_anz = type_ctx.findtext("F_UW_ANZ", default="")

                    # reset per occurrence
                    speed_min = speed_max = None
                    rpm_min = rpm_max = None
                    batt_min = batt_max = None

                    gps_time = None
                    customer_time = None

                    tnse_s = None
                    throttle_pct = None
                    tmot_c = None
                    tans_c = None

                    for uw in type_ctx.findall(".//F_UW"):
                        uw_nr = uw.findtext("F_UW_NR", default="")
                        text = uw.findtext("F_UW_TEXT", default="")
                        val_str = uw.findtext("F_UW_WERT", default="")
                        einh = uw.findtext("F_UW_EINH", default="")

                        dtc_ctx_rows.append(
                            {
                                "VIN": vin,
                                "BrandName": brand,
                                "ECUTitle": ec_title,
                                "Container": container_tag,
                                "ecuAddress": ecu_addr,
                                "dtcId": dtc_id,
                                "Id": id_field,
                                "Occurrence": occ_idx,
                                "F_UW_KM": ctx_km,
                                "F_UW_ZEIT": ctx_zeit,
                                "F_UW_ANZ": ctx_anz,
                                "F_UW_NR": uw_nr,
                                "F_UW_TEXT": text,
                                "F_UW_WERT": val_str,
                                "F_UW_EINH": einh,
                            }
                        )

                        t = (text or "").lower()
                        einh_l = (einh or "").lower()

                        if gps_time is None and ("gps" in t and "zeit" in t):
                            gps_time = val_str
                        if customer_time is None and ("kundenzeit" in t or "customer" in t):
                            customer_time = val_str

                        v = safe_float(val_str)
                        if v is None:
                            continue

                        if (("geschwindigkeit" in t or "speed" in t) and "km/h" in einh_l):
                            speed_min, speed_max = update_min_max(speed_min, speed_max, v)

                        if ("motordrehzahl" in t) or ("nmot" in t) or ("rpm" in t and "min" in einh_l):
                            rpm_min, rpm_max = update_min_max(rpm_min, rpm_max, v)

                        if ("batteriespannung" in t) or ("bordnetz" in t) or ("ub_w" in t) or (
                            "voltage" in t and "v" in einh_l
                        ):
                            batt_min, batt_max = update_min_max(batt_min, batt_max, v)

                        if tnse_s is None and ("zeit nach motorstart" in t or "tnse" in t):
                            tnse_s = v

                        if throttle_pct is None and ("fahrwertgeber" in t or "fwg" in t or "throttle" in t):
                            throttle_pct = v

                        if tmot_c is None and ("motortemperatur" in t or "tmot" in t) and ("c" in einh_l or "°" in einh_l):
                            tmot_c = v
                        if tans_c is None and ("ansauglufttemperatur" in t or "tans" in t or "intake" in t) and ("c" in einh_l or "°" in einh_l):
                            tans_c = v

                    dtc_rows.append(
                        {
                            "F_UW_ZEIT": ctx_zeit,
                            "F_UW_KM": ctx_km,
                            "F_ORT_TEXT": f_ort_text,
                            "F_UW_ANZ": ctx_anz,
                            "Speed_Min_kmh": speed_min,
                            "Speed_Max_kmh": speed_max,
                            "RPM_Min": rpm_min,
                            "RPM_Max": rpm_max,
                            "BatteryV_Min": batt_min,
                            "BatteryV_Max": batt_max,
                            "GPS_Time_UTC": gps_time,
                            "CustomerTime": customer_time,
                            "TimeSinceStart_s": tnse_s,
                            "Throttle_pct": throttle_pct,
                            "EngineTemp_C": tmot_c,
                            "IntakeTemp_C": tans_c,
                            "ECUTitle": ec_title,
                            "ECU_NAME": ecu_name,
                            "VARIANTE": variante,
                            "ECU_ADR": ecu_adr,
                            "Container": container_tag,
                            "ecuAddress": ecu_addr,
                            "dtcId": dtc_id,
                            "Id": id_field,
                            "F_ORT": f_ort,
                            "F_ART": f_art,
                            "F_VORHANDEN_TEXT": f_vorhanden_text,
                            "F_WARNUNG_TEXT": f_warnung_text,
                            "Relevance": relevance,
                            "Occurrence": occ_idx,
                            "VIN": vin,
                            "BrandName": brand,
                            "ECU_GROBNAME": ecu_grobname,
                        }
                    )

        add_dtcs("FEHLER")
        add_dtcs("INFO")

    # --- ZFSResult: BMW systeem-event log (bijv. Klemme-events, resets) ---
    for zfs in root.findall(".//ZFSResult"):
        def _t(tag: str) -> str:
            el = zfs.find(tag)
            if el is None:
                return ""
            return (el.text or "").strip()

        zfs_result_rows.append({
            "VIN": vin,
            "Index": _t("Index"),
            "STAT_DM_ADRESSE_SG": _t("STAT_DM_ADRESSE_SG"),
            "STAT_DM_MELDUNG_NR": _t("STAT_DM_MELDUNG_NR"),
            "STAT_DM_SGBD_INDEX": _t("STAT_DM_SGBD_INDEX"),
            "STAT_DM_MELDUNG_TYP": _t("STAT_DM_MELDUNG_TYP"),
            "STAT_DM_ZEITSTEMPEL": _t("STAT_DM_ZEITSTEMPEL"),
            "STAT_DM_MELDUNG_TEXT": _t("STAT_DM_MELDUNG_TEXT"),
            "STAT_DM_ACTIVE_STATE": _t("STAT_DM_ACTIVE_STATE"),
            "STAT_DM_MAPPING_ID": _t("STAT_DM_MAPPING_ID"),
            "IsCheckControlMessage": _t("IsCheckControlMessage"),
            "STAT_SYSKONTEXT_ZEITSTEMPEL_WERT": _t("STAT_SYSKONTEXT_ZEITSTEMPEL_WERT"),
            "STAT_SYSKONTEXT_ZEIT_WECKEN_WERT": _t("STAT_SYSKONTEXT_ZEIT_WECKEN_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT": _t("STAT_SYSKONTEXT_KUNDENZEIT"),
            "STAT_SYSKONTEXT_SPANNUNG_MAX_WERT": _t("STAT_SYSKONTEXT_SPANNUNG_MAX_WERT"),
            "STAT_SYSKONTEXT_SPANNUNG_MIN_WERT": _t("STAT_SYSKONTEXT_SPANNUNG_MIN_WERT"),
            "STAT_SYSKONTEXT_SPANNUNG2_MAX_WERT": _t("STAT_SYSKONTEXT_SPANNUNG2_MAX_WERT"),
            "STAT_SYSKONTEXT_SPANNUNG2_MIN_WERT": _t("STAT_SYSKONTEXT_SPANNUNG2_MIN_WERT"),
            "STAT_SYSKONTEXT_SPANNUNG_HV_SYSTEM_WERT": _t("STAT_SYSKONTEXT_SPANNUNG_HV_SYSTEM_WERT"),
            "STAT_SYSKONTEXT_GESCHWINDIGKEIT_MAX_WERT": _t("STAT_SYSKONTEXT_GESCHWINDIGKEIT_MAX_WERT"),
            "STAT_SYSKONTEXT_GESCHWINDIGKEIT_MIN_WERT": _t("STAT_SYSKONTEXT_GESCHWINDIGKEIT_MIN_WERT"),
            "STAT_SYSKONTEXT_DREHZAHL_KURBELWELLE_MAX_WERT": _t("STAT_SYSKONTEXT_DREHZAHL_KURBELWELLE_MAX_WERT"),
            "STAT_SYSKONTEXT_DREHZAHL_KURBELWELLE_MIN_WERT": _t("STAT_SYSKONTEXT_DREHZAHL_KURBELWELLE_MIN_WERT"),
            "STAT_SYSKONTEXT_TEMPERATUR_AUSSEN_WERT": _t("STAT_SYSKONTEXT_TEMPERATUR_AUSSEN_WERT"),
            "STAT_SYSKONTEXT_TEMPERATUR_MOTOR_ANTRIEB_WERT": _t("STAT_SYSKONTEXT_TEMPERATUR_MOTOR_ANTRIEB_WERT"),
            "STAT_SYSKONTEXT_KLEMMEN_BEI_FEHLER_WERT": _t("STAT_SYSKONTEXT_KLEMMEN_BEI_FEHLER_WERT"),
            "STAT_SYSKONTEXT_KLEMMEN_VOR_FEHLER_WERT": _t("STAT_SYSKONTEXT_KLEMMEN_VOR_FEHLER_WERT"),
            "STAT_SYSKONTEXT_OPSTATUS_BEI_FEHLER_WERT": _t("STAT_SYSKONTEXT_OPSTATUS_BEI_FEHLER_WERT"),
            "STAT_SYSKONTEXT_OPSTATUS_VOR_FEHLER_WERT": _t("STAT_SYSKONTEXT_OPSTATUS_VOR_FEHLER_WERT"),
            "STAT_SYSKONTEXT_FEHLERSPEICHERSPERRE_AKTIV_WERT": _t("STAT_SYSKONTEXT_FEHLERSPEICHERSPERRE_AKTIV_WERT"),
            "STAT_SYSKONTEXT_ZEIT_ERSTE_KL_R_EIN_WERT": _t("STAT_SYSKONTEXT_ZEIT_ERSTE_KL_R_EIN_WERT"),
            "STAT_SYSKONTEXT_ZEIT_ERSTE_KL_15_EIN_WERT": _t("STAT_SYSKONTEXT_ZEIT_ERSTE_KL_15_EIN_WERT"),
            "STAT_SYSKONTEXT_ZEIT_ERSTE_KL_50_EIN_WERT": _t("STAT_SYSKONTEXT_ZEIT_ERSTE_KL_50_EIN_WERT"),
            "STAT_SYSKONTEXT_ZEIT_KLEMMENWECHSEL_WERT": _t("STAT_SYSKONTEXT_ZEIT_KLEMMENWECHSEL_WERT"),
            "STAT_SYSKONTEXT_ZEIT_OPSTATUSWECHSEL_WERT": _t("STAT_SYSKONTEXT_ZEIT_OPSTATUSWECHSEL_WERT"),
            "STAT_SYSKONTEXT_ZEIT_LETZTER_PWF_WECHSEL_WERT": _t("STAT_SYSKONTEXT_ZEIT_LETZTER_PWF_WECHSEL_WERT"),
            "STAT_SYSKONTEXT_WEGSTRECKE_KILOMETER_WERT": _t("STAT_SYSKONTEXT_WEGSTRECKE_KILOMETER_WERT"),
            "STAT_SYSKONTEXT_WEGSTRECKE_INSYNC_WERT": _t("STAT_SYSKONTEXT_WEGSTRECKE_INSYNC_WERT"),
            "STAT_SYSKONTEXT_LAENGSBESCHLEUNIGUNG_WERT": _t("STAT_SYSKONTEXT_LAENGSBESCHLEUNIGUNG_WERT"),
            "STAT_SYSKONTEXT_PWF_BEI_FEHLER_WERT": _t("STAT_SYSKONTEXT_PWF_BEI_FEHLER_WERT"),
            "STAT_SYSKONTEXT_PWF_VOR_PWF_BEI_FEHLER_WERT": _t("STAT_SYSKONTEXT_PWF_VOR_PWF_BEI_FEHLER_WERT"),
            "STAT_SYSKONTEXT_SCHLSLPRFL_AKT_WERT": _t("STAT_SYSKONTEXT_SCHLSLPRFL_AKT_WERT"),
            "STAT_SYSKONTEXT_BASIS_TN_WERT": _t("STAT_SYSKONTEXT_BASIS_TN_WERT"),
            "STAT_SYSKONTEXT_FUNKT_TN_WERT": _t("STAT_SYSKONTEXT_FUNKT_TN_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_JAAR_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_JAHR_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_MONAT_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_MONAT_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_TAG_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_TAG_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_STUNDE_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_STUNDE_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_MINUTE_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_MINUTE_WERT"),
            "STAT_SYSKONTEXT_KUNDENZEIT_SEKUNDE_WERT": _t("STAT_SYSKONTEXT_KUNDENZEIT_SEKUNDE_WERT"),
        })

    ecu_df = pd.DataFrame(ecu_rows)
    dtc_df = pd.DataFrame(dtc_rows)
    ctx_df = pd.DataFrame(dtc_ctx_rows)
    zfs_result_df = pd.DataFrame(zfs_result_rows)

    dtc_df["F_UW_ZEIT_NUM"] = pd.to_numeric(dtc_df["F_UW_ZEIT"], errors="coerce")
    dtc_df["F_UW_KM_NUM"] = pd.to_numeric(dtc_df["F_UW_KM"], errors="coerce")

    if sort_mode == "km":
        dtc_df = dtc_df.sort_values(by=["F_UW_KM_NUM", "F_UW_ZEIT_NUM"], ascending=[False, False], na_position="last")
    else:
        dtc_df = dtc_df.sort_values(by=["F_UW_ZEIT_NUM", "F_UW_KM_NUM"], ascending=[False, False], na_position="last")

    dtc_df = dtc_df.drop(columns=["F_UW_ZEIT_NUM", "F_UW_KM_NUM"])

    zfs_df = (
        dtc_df.groupby(["F_UW_ZEIT", "F_UW_KM"], dropna=False)
        .agg(
            Speed_Min_kmh=("Speed_Min_kmh", "min"),
            Speed_Max_kmh=("Speed_Max_kmh", "max"),
            RPM_Min=("RPM_Min", "min"),
            RPM_Max=("RPM_Max", "max"),
            BatteryV_Min=("BatteryV_Min", "min"),
            BatteryV_Max=("BatteryV_Max", "max"),
            GPS_Time_UTC=("GPS_Time_UTC", "first"),
            CustomerTime=("CustomerTime", "first"),
            DTC_Count=("dtcId", "count"),
        )
        .reset_index()
    )

    # --- ZFS tijdcorrectie: offset-analyse op basis van ZFSResult kalibratiedata ---
    if not zfs_result_df.empty:

        def _kz_to_unix(s: str):
            try:
                dt = datetime.fromisoformat(str(s).strip())
                return (dt - datetime(1970, 1, 1)).total_seconds()
            except Exception:
                return None

        zfs_result_df["_zfs_num"] = pd.to_numeric(
            zfs_result_df["STAT_DM_ZEITSTEMPEL"], errors="coerce"
        )
        zfs_result_df["_kz_unix"] = zfs_result_df["STAT_SYSKONTEXT_KUNDENZEIT"].apply(_kz_to_unix)
        zfs_result_df["ZFS_Offset_s"] = (
            zfs_result_df["_kz_unix"] - zfs_result_df["_zfs_num"]
        ).round().astype("Int64")

        calib = (
            zfs_result_df.dropna(subset=["_zfs_num", "_kz_unix", "ZFS_Offset_s"])
            .sort_values("_zfs_num")
            .reset_index(drop=True)
        )

        if not calib.empty:
            offsets = calib["ZFS_Offset_s"].astype(float)
            delta = offsets.diff().fillna(0)
            # Alleen positieve sprongen ≥30 s tellen als onderbreking (negatief = meetnoise)
            calib["Onderbreking_voor_s"] = delta.where(delta >= 30, 0).astype(int)

            zfs_to_interr = dict(zip(calib["_zfs_num"], calib["Onderbreking_voor_s"]))
            zfs_result_df["Onderbreking_voor_s"] = (
                zfs_result_df["_zfs_num"].map(zfs_to_interr).fillna(0).astype(int)
            )

            calib_zfs_arr = calib["_zfs_num"].values
            calib_off_arr = calib["ZFS_Offset_s"].astype(float).values

            def _zfs_to_wallclock(zfs_val: str) -> str:
                try:
                    v = float(zfs_val)
                except Exception:
                    return ""
                # Gebruik de meest recente kalibratiepunt op of vóór dit ZFS-moment
                mask = calib_zfs_arr <= v
                idx = int(mask.nonzero()[0][-1]) if mask.any() else 0
                corrected = v + calib_off_arr[idx]
                return (datetime(1970, 1, 1) + timedelta(seconds=corrected)).strftime(
                    "%Y-%m-%d %H:%M:%S"
                )

            if "F_UW_ZEIT" in dtc_df.columns and not dtc_df.empty:
                col_pos = dtc_df.columns.get_loc("F_UW_ZEIT") + 1
                dtc_df.insert(
                    col_pos, "Tijdstip_Klant", dtc_df["F_UW_ZEIT"].apply(_zfs_to_wallclock)
                )

        zfs_result_df = zfs_result_df.drop(columns=["_zfs_num", "_kz_unix"], errors="ignore")

    summary = {
        "VIN": vin,
        "BrandName": brand,
        "Script_Version": SCRIPT_VERSION,
        "Designer": DESIGNER,
        "Department": DEPARTMENT,
        "TimeReference": "F_UW_ZEIT is a relative vehicle time counter (ZFS), not epoch/UTC.",
    }
    for k, v in (meta or {}).items():
        summary[f"Meta_{k}"] = v

    summary_df = pd.DataFrame(list(summary.items()), columns=["Key", "Value"])

    out_path = Path(out_excel_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Voorblad", index=False)
        ecu_df.to_excel(writer, sheet_name="ECUs", index=False)
        dtc_df.to_excel(writer, sheet_name="DTCs", index=False)
        ctx_df.to_excel(writer, sheet_name="DTC_Context", index=False)
        zfs_df.to_excel(writer, sheet_name="ZFS_Seconds", index=False)
        # Kolomvolgorde en sortering ZFS_Resultaten
        _zfs_priority = [
            "STAT_SYSKONTEXT_KUNDENZEIT",
            "STAT_DM_MELDUNG_TEXT",
            "STAT_SYSKONTEXT_GESCHWINDIGKEIT_MAX_WERT",
            "STAT_SYSKONTEXT_GESCHWINDIGKEIT_MIN_WERT",
            "ZFS_Offset_s",
            "Onderbreking_voor_s",
        ]
        _zfs_rest = [c for c in zfs_result_df.columns if c not in _zfs_priority]
        _zfs_cols = [c for c in _zfs_priority if c in zfs_result_df.columns] + _zfs_rest
        zfs_out = zfs_result_df[_zfs_cols].copy()
        if "STAT_DM_ZEITSTEMPEL" in zfs_out.columns:
            zfs_out["_sort"] = pd.to_numeric(zfs_out["STAT_DM_ZEITSTEMPEL"], errors="coerce")
            zfs_out = zfs_out.sort_values("_sort", ascending=False, na_position="last").drop(columns=["_sort"])
        zfs_out.to_excel(writer, sheet_name="ZFS_Resultaten", index=False)

        ws = writer.book["DTCs"]
        if "F_ORT_TEXT" in dtc_df.columns:
            col_idx = dtc_df.columns.get_loc("F_ORT_TEXT") + 1
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = int(f_ort_width)

            if wrap_f_ort_text:
                alignment = Alignment(wrap_text=True, vertical="top")
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = alignment


class BMWXMLApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(f"BMW XML Verwerker {SCRIPT_VERSION} | {DESIGNER} | @{DEPARTMENT}")
        self.root.resizable(False, False)

        self.meta_path = tk.StringVar()
        self.trans_path = tk.StringVar()
        self.status_var = tk.StringVar(value="Selecteer beide XML-bestanden en klik op Verwerken.")

        self._build_ui()

    # ------------------------------------------------------------------
    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 5}

        frame = tk.Frame(self.root, padx=18, pady=14)
        frame.pack(fill="both", expand=True)

        # --- Header met logo en versie ---
        header = tk.Frame(frame)
        header.grid(row=0, column=0, columnspan=3, sticky="ew", pady=(0, 10))

        try:
            self.logo_image = tk.PhotoImage(file=str(resource_path("BMW_XML.png")))
            max_h = 72
            if self.logo_image.height() > max_h:
                factor = max(1, int(self.logo_image.height() / max_h))
                self.logo_image = self.logo_image.subsample(factor, factor)
            tk.Label(header, image=self.logo_image).pack(side="left", padx=(0, 12))
        except Exception:
            tk.Label(header, text="BMW", font=("Segoe UI", 18, "bold"), fg="#005a9e").pack(side="left", padx=(0, 12))

        title_box = tk.Frame(header)
        title_box.pack(side="left", fill="x", expand=True)
        tk.Label(title_box, text="BMW XML Verwerker", font=("Segoe UI", 15, "bold"), anchor="w").pack(anchor="w")
        tk.Label(
            title_box,
            text=f"Auteur: {DESIGNER}  |  Afdeling: @{DEPARTMENT}  |  Versie: {SCRIPT_VERSION}",
            font=("Segoe UI", 9),
            fg="#444444",
            anchor="w",
        ).pack(anchor="w")

        # --- RG_META ---
        tk.Label(frame, text="RG_META XML:", anchor="w").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(frame, textvariable=self.meta_path, width=58).grid(row=1, column=1, **pad)
        tk.Button(frame, text="Bladeren...", width=12, command=self._browse_meta).grid(row=1, column=2, **pad)

        # --- RG_TRANS ---
        tk.Label(frame, text="RG_TRANS XML:", anchor="w").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(frame, textvariable=self.trans_path, width=58).grid(row=2, column=1, **pad)
        tk.Button(frame, text="Bladeren...", width=12, command=self._browse_trans).grid(row=2, column=2, **pad)

        # --- Verwerken knop ---
        self.btn_run = tk.Button(
            frame, text="Verwerken", width=18,
            bg="#005a9e", fg="white", activebackground="#003f6e", activeforeground="white",
            font=("Segoe UI", 10, "bold"),
            command=self._run,
        )
        self.btn_run.grid(row=3, column=0, columnspan=3, pady=(14, 6))

        # --- Voortgangsbalk ---
        self.progress = ttk.Progressbar(frame, mode="indeterminate", length=460)
        self.progress.grid(row=4, column=0, columnspan=3, pady=(0, 8))

        # --- Statusregel ---
        tk.Label(frame, textvariable=self.status_var, wraplength=500, justify="left",
                 anchor="w").grid(row=5, column=0, columnspan=3, sticky="w")

    # ------------------------------------------------------------------
    def _browse_meta(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecteer RG_META XML",
            filetypes=[("XML-bestanden", "*.xml"), ("Alle bestanden", "*.*")],
        )
        if not path:
            return
        self.meta_path.set(path)
        # Probeer RG_TRANS automatisch in te vullen
        if not self.trans_path.get():
            for f in Path(path).parent.glob("*RG_TRANS*.xml"):
                self.trans_path.set(str(f))
                break

    def _browse_trans(self) -> None:
        path = filedialog.askopenfilename(
            title="Selecteer RG_TRANS XML",
            filetypes=[("XML-bestanden", "*.xml"), ("Alle bestanden", "*.*")],
        )
        if not path:
            return
        self.trans_path.set(path)
        # Probeer RG_META automatisch in te vullen
        if not self.meta_path.get():
            for f in Path(path).parent.glob("*RG_META*.xml"):
                self.meta_path.set(str(f))
                break

    # ------------------------------------------------------------------
    def _run(self) -> None:
        meta = self.meta_path.get().strip()
        trans = self.trans_path.get().strip()

        if not meta or not trans:
            messagebox.showwarning(
                "Ontbrekende bestanden",
                "Selecteer zowel het RG_META als het RG_TRANS XML-bestand.",
            )
            return

        for label, p in (("RG_META", meta), ("RG_TRANS", trans)):
            if not Path(p).exists():
                messagebox.showerror("Bestand niet gevonden", f"{label} niet gevonden:\n{p}")
                return

        self.btn_run.config(state="disabled")
        self.progress.start(10)
        self.status_var.set("Bezig met verwerken…")

        threading.Thread(target=self._process, args=(meta, trans), daemon=True).start()

    def _process(self, meta: str, trans: str) -> None:
        try:
            out_dir = Path(trans).parent
            vin = extract_vin(trans) or extract_vin(meta) or "UNKNOWNVIN"
            out_excel = out_dir / f"BMW_XML_{vin}.xlsx"

            parse_vehicle_xml_to_excel(
                trans_xml_path=trans,
                meta_xml_path=meta,
                out_excel_path=str(out_excel),
            )
            self.root.after(0, self._done, str(out_excel), None)
        except Exception as e:
            tb = traceback.format_exc()
            log_path = Path(trans).parent / "BMW_XML_debug.log"
            write_log(Path(trans).parent, tb)
            self.root.after(0, self._done, None, f"{e}\n\nLogbestand: {log_path}")

    def _done(self, out_path: str | None, error: str | None) -> None:
        self.progress.stop()
        self.btn_run.config(state="normal")

        if error:
            self.status_var.set("Fout opgetreden — zie logbestand.")
            messagebox.showerror("Fout", error)
        else:
            self.status_var.set(f"Klaar!  Excel opgeslagen in:\n{out_path}")
            messagebox.showinfo("Klaar", f"Excel gemaakt:\n{out_path}")


# V4 enhanced parser override: keeps the existing GUI, but uses the richer workbook export.
try:
    from bmw_xml_enhanced import parse_vehicle_xml_to_excel, extract_vin, SCRIPT_VERSION, DESIGNER, DEPARTMENT
except Exception:
    pass
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = BMWXMLApp(root)

    # Drag-and-drop of command-line argumenten vooraf invullen
    args = sys.argv[1:]
    if args:
        meta_pre, trans_pre = pick_paths_from_argv(args)
        if meta_pre:
            app.meta_path.set(meta_pre)
        if trans_pre:
            app.trans_path.set(trans_pre)

    root.mainloop()


