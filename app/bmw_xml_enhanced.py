import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCRIPT_VERSION = "V4.1"
DESIGNER = "MvR"
DEPARTMENT = "tdOob"


def strip_namespaces(root):
    for el in root.iter():
        if "}" in el.tag:
            el.tag = el.tag.split("}", 1)[1]
        attrs = {}
        for k, v in el.attrib.items():
            attrs[k.split("}", 1)[-1] if "}" in k else k] = v
        el.attrib.clear(); el.attrib.update(attrs)


def parse_xml(path):
    if not path or not Path(path).exists():
        return None
    root = ET.parse(path).getroot()
    strip_namespaces(root)
    return root


def t(el, default=""):
    return (el.text or default).strip() if el is not None else default


def sf(value):
    try:
        return float(str(value).replace(",", "."))
    except Exception:
        return None


def mm(lo, hi, value):
    if value is None:
        return lo, hi
    return value if lo is None or value < lo else lo, value if hi is None or value > hi else hi


def dur(start, end):
    try:
        if not start or not end or str(end).startswith("0001-"):
            return None
        return round((datetime.fromisoformat(end) - datetime.fromisoformat(start)).total_seconds(), 1)
    except Exception:
        return None


def extract_vin(xml_path):
    root = parse_xml(xml_path)
    if root is None:
        return ""
    return (root.findtext("VCI/VIN") or root.findtext("VIN17") or root.findtext(".//Vin17") or "").strip()


def first_file(folder, patterns):
    for pattern in patterns:
        hits = sorted(Path(folder).glob(pattern))
        if hits:
            return hits[0]
    return None


def related(trans_path, meta_path):
    folder = Path(trans_path or meta_path).parent
    return {
        "prg": first_file(folder, ["*RG_PRG*.xml"]),
        "fstdat": first_file(folder, ["*.fstdat"]),
        "behdat": first_file(folder, ["*.behdat"]),
    }


def meta_rows(meta_path):
    root = parse_xml(meta_path)
    if root is None:
        return []
    wanted = {
        "VIN17": "VIN17", "Baureihe": "BasicFeatures/Baureihe", "Ereihe": "BasicFeatures/Ereihe",
        "Karosserie": "BasicFeatures/Karosserie", "VerkaufsBezeichnung": "BasicFeatures/VerkaufsBezeichnung",
        "Motor": "BasicFeatures/Motor", "Getriebe": "BasicFeatures/Getriebe", "Modelljaar": "BasicFeatures/Modelljahr",
        "Modellmaand": "BasicFeatures/Modellmonat", "Marke": "BasicFeatures/Marke", "TypeCode": "BasicFeatures/TypeCode",
        "StartDate": "StartDate", "EndDate": "EndDate", "WorkState": "WorkState", "VehicleCommunication": "VehicleCommunication",
        "ComputerName": "ComputerName", "UserName": "UserName", "IstaCaseId": "IstaCaseId",
        "DateOfFastaRead": "DateOfFastaRead", "DistanceOfFastaRead": "DistanceOfFastaRead",
        "DealerNumber": "DealerNumber", "distributionPartnerNumber": "distributionPartnerNumber", "outletNumber": "outletNumber",
    }
    return [(k, t(root.find(p))) for k, p in wanted.items()]


def dtc_base(vin, brand, ecu, container, dtc, status):
    row = {
        "F_UW_ZEIT": "", "F_UW_KM": "", "F_ORT_TEXT": t(dtc.find("F_ORT_TEXT")), "F_UW_ANZ": "",
        "Speed_Min_kmh": None, "Speed_Max_kmh": None, "RPM_Min": None, "RPM_Max": None,
        "BatteryV_Min": None, "BatteryV_Max": None, "GPS_Time_UTC": None, "CustomerTime": None,
        "TimeSinceStart_s": None, "Throttle_pct": None, "EngineTemp_C": None, "IntakeTemp_C": None,
        "ECUTitle": ecu["ECUTitle"], "ECU_NAME": ecu["ECU_NAME"], "VARIANTE": ecu["VARIANTE"], "ECU_ADR": ecu["ECU_ADR"],
        "Container": container, "ecuAddress": t(dtc.find("ecuAddress")), "dtcId": t(dtc.find("dtcId")), "Id": t(dtc.find("Id")),
        "F_ORT": t(dtc.find("F_ORT")), "F_ART": t(dtc.find("F_ART")), "F_VORHANDEN_TEXT": t(dtc.find("F_VORHANDEN_TEXT")),
        "F_WARNUNG_TEXT": t(dtc.find("F_WARNUNG_TEXT")), "Relevance": t(dtc.find("Relevance")), "Occurrence": None,
        "Context_Status": status, "F_HEX_CODE": t(dtc.find("F_HEX_CODE")), "F_READY_TEXT": t(dtc.find("F_READY_TEXT")),
        "F_FEHLERKLASSE_TEXT": t(dtc.find("F_FEHLERKLASSE_TEXT")), "F_SAE_CODE": t(dtc.find("F_SAE_CODE")), "F_PCODE": t(dtc.find("F_PCODE")),
        "FaultClass": t(dtc.find("FaultClass")), "FaultGroup": t(dtc.find("FaultGroup")), "StateWarningLight": t(dtc.find("StateWarningLight")),
        "VIN": vin, "BrandName": brand, "ECU_GROBNAME": ecu["ECU_GROBNAME"],
    }
    return row


def parse_trans(root, vin, brand):
    ecu_rows, svk_rows, dtc_rows, ctx_rows, zfs_rows = [], [], [], [], []
    for node in root.findall("ECU/ECU"):
        ecu = {"ECUTitle": t(node.find("ECUTitle")), "VARIANTE": t(node.find("VARIANTE")), "ECU_GROBNAME": t(node.find("ECU_GROBNAME")), "ECU_NAME": t(node.find("ECU_NAME")), "ECU_ADR": t(node.find("ECU_ADR"))}
        svk = node.find("SVK")
        xwe_node = svk.find("XWE_SGBMID") if svk is not None else None
        xwe_parts = [t(x) for x in xwe_node.findall("string")] if xwe_node is not None else []
        xwe = ", ".join([x for x in xwe_parts if x])
        prog_date = t(svk.find("PROG_DATUM") if svk is not None else None)
        prog_km = t(svk.find("PROG_KM") if svk is not None else None)
        ecu_rows.append({"VIN": vin, "BrandName": brand, "Generation": t(node.find("Generation")), **ecu, "ECU_SGBD": t(node.find("ECU_SGBD")), "ECU_GRUPPE": t(node.find("ECU_GRUPPE")), "BUS": t(node.find("BUS")), "SubBUS": t(node.find("SubBUS")), "DiagProtocoll": t(node.find("DiagProtocoll")), "ID_LIEF_NR": t(node.find("ID_LIEF_NR")), "ID_LIEF_TEXT": t(node.find("ID_LIEF_TEXT")), "ID_DATUM": t(node.find("ID_DATUM")), "ID_SW_NR": t(node.find("ID_SW_NR")), "ID_SGBD_INDEX": t(node.find("ID_SGBD_INDEX")), "ID_SG_ADR": t(node.find("ID_SG_ADR")), "SERIENNUMMER": t(node.find("SERIENNUMMER")), "F_ANZ": t(node.find("F_ANZ")), "I_ANZ": t(node.find("I_ANZ")), "XWE_SGBMID": xwe, "PROG_DATUM": prog_date, "PROG_KM": prog_km, "COMMUNICATION_SUCCESSFULLY": t(node.find("COMMUNICATION_SUCCESSFULLY")), "IDENT_SUCCESSFULLY": t(node.find("IDENT_SUCCESSFULLY")), "FS_SUCCESSFULLY": t(node.find("FS_SUCCESSFULLY")), "IS_SUCCESSFULLY": t(node.find("IS_SUCCESSFULLY")), "SERIAL_SUCCESSFULLY": t(node.find("SERIAL_SUCCESSFULLY")), "SVK_SUCCESSFULLY": t(node.find("SVK_SUCCESSFULLY")), "HW_REF_STATUS": t(node.find("HW_REF_STATUS")), "SWTStatus": t(node.find("SWTStatus"))})
        for part in [p.strip() for p in xwe.split(",") if p.strip()]:
            svk_rows.append({"VIN": vin, "ECUTitle": ecu["ECUTitle"], "ECU_GROBNAME": ecu["ECU_GROBNAME"], "ECU_ADR": ecu["ECU_ADR"], "SGBM_ID": part, "SGBM_Type": part.split("-", 1)[0] if "-" in part else "", "PROG_DATUM": prog_date, "PROG_KM": prog_km})
        for container_name in ["FEHLER", "INFO"]:
            container = node.find(container_name)
            if container is None:
                continue
            for dtc in container.findall("DTC"):
                dtc_context = dtc.find("DTCContext")
                type_contexts = dtc_context.findall("typeDTCContext") if dtc_context is not None else []
                if not type_contexts:
                    dtc_rows.append(dtc_base(vin, brand, ecu, container_name, dtc, "Empty DTCContext" if dtc_context is not None else "No DTCContext"))
                    continue
                for occurrence, type_ctx in enumerate(type_contexts, start=1):
                    row = dtc_base(vin, brand, ecu, container_name, dtc, "Has typeDTCContext")
                    row.update({"F_UW_KM": t(type_ctx.find("F_UW_KM")), "F_UW_ZEIT": t(type_ctx.find("F_UW_ZEIT")), "F_UW_ANZ": t(type_ctx.find("F_UW_ANZ")), "Occurrence": occurrence})
                    speed_min = speed_max = rpm_min = rpm_max = batt_min = batt_max = None
                    for uw in type_ctx.findall(".//F_UW"):
                        text, value, unit = t(uw.find("F_UW_TEXT")), t(uw.find("F_UW_WERT")), t(uw.find("F_UW_EINH"))
                        ctx_rows.append({"VIN": vin, "BrandName": brand, "ECUTitle": ecu["ECUTitle"], "Container": container_name, "ecuAddress": t(dtc.find("ecuAddress")), "dtcId": t(dtc.find("dtcId")), "Id": t(dtc.find("Id")), "Occurrence": occurrence, "F_UW_KM": row["F_UW_KM"], "F_UW_ZEIT": row["F_UW_ZEIT"], "F_UW_ANZ": row["F_UW_ANZ"], "F_UW_NR": t(uw.find("F_UW_NR")), "F_UW_NAME": t(uw.find("F_UW_NAME")), "F_UW_TYP": t(uw.find("F_UW_TYP")), "F_UW_TEXT": text, "F_UW_WERT": value, "F_UW_EINH": unit})
                        low, low_unit, num = text.lower(), unit.lower(), sf(value)
                        if row["GPS_Time_UTC"] is None and "gps" in low and "zeit" in low: row["GPS_Time_UTC"] = value
                        if row["CustomerTime"] is None and ("kundenzeit" in low or "customer" in low): row["CustomerTime"] = value
                        if num is None: continue
                        if ("geschwindigkeit" in low or "speed" in low) and "km/h" in low_unit: speed_min, speed_max = mm(speed_min, speed_max, num)
                        if "motordrehzahl" in low or "nmot" in low or ("rpm" in low and "min" in low_unit): rpm_min, rpm_max = mm(rpm_min, rpm_max, num)
                        if "batteriespannung" in low or "bordnetz" in low or "ub_w" in low or ("voltage" in low and "v" in low_unit): batt_min, batt_max = mm(batt_min, batt_max, num)
                        if row["TimeSinceStart_s"] is None and ("zeit nach motorstart" in low or "tnse" in low): row["TimeSinceStart_s"] = num
                        if row["Throttle_pct"] is None and ("fahrwertgeber" in low or "fwg" in low or "throttle" in low): row["Throttle_pct"] = num
                        if row["EngineTemp_C"] is None and ("motortemperatur" in low or "tmot" in low) and ("c" in low_unit or "°" in low_unit): row["EngineTemp_C"] = num
                        if row["IntakeTemp_C"] is None and ("ansauglufttemperatur" in low or "tans" in low or "intake" in low) and ("c" in low_unit or "°" in low_unit): row["IntakeTemp_C"] = num
                    row.update({"Speed_Min_kmh": speed_min, "Speed_Max_kmh": speed_max, "RPM_Min": rpm_min, "RPM_Max": rpm_max, "BatteryV_Min": batt_min, "BatteryV_Max": batt_max})
                    dtc_rows.append(row)
    for zfs in root.findall(".//ZFSResult"):
        row = {"VIN": vin}
        for child in list(zfs):
            if len(list(child)) == 0:
                row[child.tag] = t(child)
        zfs_rows.append(row)
    return pd.DataFrame(ecu_rows), pd.DataFrame(svk_rows), pd.DataFrame(dtc_rows), pd.DataFrame(ctx_rows), pd.DataFrame(zfs_rows)


def parse_prg(path, vin):
    root = parse_xml(path)
    if root is None:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    ecus, buses, sgbms = [], [], []
    ecu_nodes = [n for n in root.findall(".//anyType") if "PsdzEcu" in n.attrib.get("type", "")]
    sgbm_nodes = [n for n in root.findall(".//anyType") if "PsdzSgbmId" in n.attrib.get("type", "")]
    for i, ecu in enumerate(ecu_nodes, 1):
        row = {"VIN": vin, "PRG_ECU_Index": i}
        for tag in ["BaseVariant", "BnTnName", "DiagnosticBus", "EcuVariant", "GatewayDiagAddr", "PrimaryKey", "DiagnosisAddress", "IsSmartActuator", "Offset", "SerialNumber"]:
            row[tag] = t(ecu.find(tag))
        for prefix, node_name in [("Detail", "EcuDetailInfo"), ("Status", "EcuStatusInfo"), ("Pdx", "PsdzEcuPdxInfo")]:
            node = ecu.find(node_name)
            if node is not None:
                for child in list(node):
                    if len(list(child)) == 0:
                        row[f"{prefix}_{child.tag}"] = t(child)
        ecus.append(row)
        for bus in ecu.findall(".//PsdzBus"):
            buses.append({"VIN": vin, "PRG_ECU_Index": i, "BaseVariant": row.get("BaseVariant", ""), "PsdzBus": t(bus)})
    for sgbm in sgbm_nodes:
        sgbms.append({"VIN": vin, "Id": t(sgbm.find("Id")), "HexString": t(sgbm.find("HexString")), "IdAsLong": t(sgbm.find("IdAsLong")), "ProcessClass": t(sgbm.find("ProcessClass")), "MainVersion": t(sgbm.find("MainVersion")), "SubVersion": t(sgbm.find("SubVersion")), "PatchVersion": t(sgbm.find("PatchVersion")), "SGBMIDVersion": t(sgbm.find("SGBMIDVersion"))})
    return pd.DataFrame(ecus), pd.DataFrame(buses), pd.DataFrame(sgbms)

def parse_fasta(path, vin):
    root = parse_xml(path)
    if root is None: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    vehicle, equipment, ecus, funcs = [], [], [], []
    vt, vo = root.find(".//VehicleTest"), root.find(".//VehicleOrder")
    if vt is not None: vehicle.append({"VIN": vin, "Type": vt.attrib.get("Type", ""), "ResultIO": vt.attrib.get("ResultIO", ""), "StartTime": vt.attrib.get("StartTime", ""), "EndTime": vt.attrib.get("EndTime", ""), "Duration_s": dur(vt.attrib.get("StartTime", ""), vt.attrib.get("EndTime", "")), "ILevelBuild": t(vt.find("ILevelBuild")), "ILevelHO": t(vt.find("ILevelHO")), "Distance": t(vt.find("Distance"))})
    if vo is not None:
        base = {"VIN": vin, "VOValue": t(vo.find("VOValue")), "VOeSerie": t(vo.find("VOeSerie")), "Type": t(vo.find("Type")), "TimeCriteria": t(vo.find("TimeCriteria")), "PaintCode": t(vo.find("PaintCode")), "CushionCode": t(vo.find("CushionCode"))}
        for eq in vo.findall(".//Equipment"): equipment.append({**base, "CodeType": "Equipment", "Code": t(eq)})
        for ew in vo.findall(".//EWord"): equipment.append({**base, "CodeType": "EWord", "Code": t(ew)})
    for ecu in root.findall(".//Ecu"):
        ident = ecu.find(".//EcuIdent")
        ecus.append({"VIN": vin, "EcuName": ecu.attrib.get("EcuName", ""), "EcuResult": ecu.attrib.get("EcuResult", ""), "StartTime": ecu.attrib.get("StartTime", ""), "EndTime": ecu.attrib.get("EndTime", ""), "Duration_s": dur(ecu.attrib.get("StartTime", ""), ecu.attrib.get("EndTime", "")), "EcuDFVariant": t(ident.find("EcuDFVariant") if ident is not None else None), "EcuVariant": t(ident.find("EcuVariant") if ident is not None else None), "DiagnosticAddress": t(ident.find("DiagnosticAddress") if ident is not None else None), "Function_Count": len(ecu.findall(".//EcuFunction")), "Record_Count": len(ecu.findall(".//Record")), "Result_Count": len(ecu.findall(".//Result"))})
        for func in ecu.findall(".//EcuFunction"):
            funcs.append({"VIN": vin, "EcuName": ecu.attrib.get("EcuName", ""), "JobName": func.attrib.get("JobName", ""), "JobStatus": func.attrib.get("JobStatus", ""), "Argument": " | ".join(t(a) or a.attrib.get("Name", "") for a in func.findall("Argument")), "Record_Count": len(func.findall("Record")), "Result_Count": len(func.findall(".//Result")), "Result_Names": ", ".join(sorted({r.attrib.get("Name", "") for r in func.findall(".//Result") if r.attrib.get("Name")}))[:500]})
    return pd.DataFrame(vehicle), pd.DataFrame(equipment), pd.DataFrame(ecus), pd.DataFrame(funcs)


def parse_beh(path, vin):
    root = parse_xml(path)
    if root is None: return pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    backend, tests, actions, events, devices = [], [], [], [], []
    for br in root.findall(".//BackendRequest"):
        st = br.find("Status")
        backend.append({"VIN": vin, "UseCase": br.attrib.get("UseCase", ""), "UseCaseVersion": br.attrib.get("UseCaseVersion", ""), "NumberOfCalls": t(br.find("NumberOfCalls")), "StatusCode": t(st.find("StatusCode") if st is not None else None), "StatusDescription": t(st.find("StatusDescription") if st is not None else None), "Object": t(st.find("Object") if st is not None else None), "Element_Count": len(br.findall(".//Element"))})
    for tm in root.findall(".//TestModule"):
        titles = {x.attrib.get("Language", ""): t(x) for x in tm.findall("Title")}
        tests.append({"VIN": vin, "Identifier": tm.attrib.get("Identifier", ""), "Title_NL": titles.get("nl-NL", ""), "Title_EN": titles.get("en-GB", ""), "Result": tm.attrib.get("Result", ""), "CollectiveResult": tm.attrib.get("CollectiveResult", ""), "Source": tm.attrib.get("Source", ""), "StartTime": tm.attrib.get("StartTime", ""), "EndTime": tm.attrib.get("EndTime", ""), "Duration_s": dur(tm.attrib.get("StartTime", ""), tm.attrib.get("EndTime", "")), "Step_Count": len(tm.findall(".//ModuleStep")), "SubModule_Count": len(tm.findall(".//SubModule"))})
    for act in root.findall(".//Action"):
        sc = act.find("ServiceCodeEnter")
        actions.append({"VIN": vin, "StartTime": act.attrib.get("StartTime", ""), "EndTime": act.attrib.get("EndTime", ""), "Duration_s": dur(act.attrib.get("StartTime", ""), act.attrib.get("EndTime", "")), "Result": act.attrib.get("Result", ""), "LayoutGroup": act.attrib.get("LayoutGroup", ""), "ServiceCodeName": t(sc.find("Name") if sc is not None else None), "ServiceCodeValue": t(sc.find("Value") if sc is not None else None), "HasEcuCommunication": act.find(".//EcuCommunication") is not None, "HasInfoLog": act.find(".//InfoLog") is not None})
    for ev in root.findall(".//Event"):
        child = list(ev)[0] if list(ev) else None
        row = {"VIN": vin, "StartTime": ev.attrib.get("StartTime", ""), "LayoutGroup": ev.attrib.get("LayoutGroup", ""), "EventType": child.tag if child is not None else ""}
        if child is not None: row.update(child.attrib)
        events.append(row)
    for dev in root.findall(".//Device"):
        row = {"VIN": vin, **dev.attrib}
        for tag in ["DeviceType", "Name", "DeviceVersion", "SerialNumber", "IpAddress", "MacAddress", "SoftwareVersion", "ConnectionType", "ReasonDisconnect"]: row[tag] = t(dev.find(tag))
        devices.append(row)
    return pd.DataFrame(backend), pd.DataFrame(tests), pd.DataFrame(actions), pd.DataFrame(events), pd.DataFrame(devices)


def add_table(ws, name):
    from openpyxl.worksheet.table import Table, TableStyleInfo
    if ws.max_row < 2: return
    tab = Table(displayName=name[:255], ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 8
        for cell in ws[letter][:min(ws.max_row, 200)]:
            if cell.value is not None: width = max(width, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(width + 2, 60)


def documentation_df(rel, trans_xml_path, meta_xml_path):
    source_rows = [
        ("Algemeen", "Auteur", DESIGNER),
        ("Algemeen", "Afdeling", DEPARTMENT),
        ("Algemeen", "Versie", SCRIPT_VERSION),
        ("Gebruikte bestanden", "RG_TRANS", Path(trans_xml_path).name if trans_xml_path else "", "Hoofdbron voor ECU-overzicht, DTC's, DTC-context, ZFS-resultaten en systeemcontext."),
        ("Gebruikte bestanden", "RG_META", Path(meta_xml_path).name if meta_xml_path else "", "Metadata van de diagnose-/transactiesessie, voertuigbasisgegevens, dealer, gebruiker en FASTA-leesmoment."),
        ("Gebruikte bestanden", "RG_PRG", rel["prg"].name if rel.get("prg") else "Niet gevonden", "Programmeersessiegegevens: ECU-detailinformatie, busdata, SGBM/software-identificaties en security/certificaatflags."),
        ("Gebruikte bestanden", "FSTDAT", rel["fstdat"].name if rel.get("fstdat") else "Niet gevonden", "FASTA/voertuigtestdata: vehicle order, uitrustingscodes, I-levels, ECU-testresultaten en uitgevoerde ECU-functies."),
        ("Gebruikte bestanden", "BEHDAT", rel["behdat"].name if rel.get("behdat") else "Niet gevonden", "Sessielog/protocoldata: acties, testmodules, backend-aanvragen, events, batterijmelding en ICOM/device-informatie."),
    ]
    sheet_rows = [
        ("Tabbladen", "Dashboard", "", "KPI-overzicht, DTC's per ECU, top foutteksten en DTC-tijdlijn."),
        ("Tabbladen", "Voorblad", "", "Samenvatting van voertuig, scriptversie, gebruikte bestanden, DTC-tellingen en metadata uit RG_META."),
        ("Tabbladen", "ECUs", "", "ECU-overzicht uit RG_TRANS met communicatie-/identificatiestatus, adressen, leverancier en basis-SVK-informatie."),
        ("Tabbladen", "DTCs", "", "Hoofdtable met foutcodes en info-DTC's. Bevat ook DTC's met lege context, plus verrijkte velden zoals F_HEX_CODE, F_READY_TEXT en foutklasse."),
        ("Tabbladen", "DTC_Context", "", "Ruwe contextregels per DTC occurrence: F_UW-nummer, naam/type, tekst, waarde en eenheid."),
        ("Tabbladen", "ZFS_Seconds", "", "Groepering van DTC's per voertuig-ZFS-tijd en kilometerstand met min/max waarden voor snelheid, RPM en batterijspanning."),
        ("Tabbladen", "ZFS_Resultaten", "", "Ruwe ZFSResult eventlog uit RG_TRANS inclusief klanttijd, meldingstekst, spanning, snelheid, klemstatus en offsetanalyse."),
        ("Tabbladen", "SVK_Software", "", "Software-/SGBM-ID's per ECU uit RG_TRANS SVK-blokken, opgesplitst naar type zoals HWEL, BTLD, SWFL en CAFD."),
        ("Tabbladen", "PRG_ECUs", "", "ECU-detailrecords uit RG_PRG met Psdz ECU-eigenschappen en security/certificaatflags."),
        ("Tabbladen", "PRG_Bus", "", "Buskoppelingen per PRG ECU-record."),
        ("Tabbladen", "PRG_SGBM", "", "Alle SGBM/software-identificaties uit RG_PRG met procesklasse en versies."),
        ("Tabbladen", "FASTA_Vehicle", "", "Voertuigtest-samenvatting uit FSTDAT: testtype, resultaat, duur, I-levels en kilometerstand."),
        ("Tabbladen", "FASTA_Equipment", "", "Vehicle order en uitrustings-/E-word codes uit FSTDAT."),
        ("Tabbladen", "FASTA_ECUs", "", "FASTA ECU-lijst met testresultaat, diagnostisch adres en aantallen functies/records/results."),
        ("Tabbladen", "FASTA_Functions", "", "Uitgevoerde ECU-functies/jobs uit FSTDAT met jobstatus en resultaatvelden."),
        ("Tabbladen", "BEH_Backend", "", "Backend-aanvragen uit BEHDAT met use-case, statuscode, omschrijving en aantal elementen."),
        ("Tabbladen", "BEH_Testmodules", "", "Uitgevoerde testmodules uit BEHDAT met resultaat, collective result, bron, duur en stappen."),
        ("Tabbladen", "BEH_Actions", "", "Sessietijdlijn van acties uit BEHDAT inclusief servicecodes en indicatie of ECU-communicatie of infolog aanwezig was."),
        ("Tabbladen", "BEH_Events", "", "Events uit BEHDAT, zoals batterijwaarschuwingen en schermwissels."),
        ("Tabbladen", "BEH_Device", "", "Gebruikte interface/device-informatie, bijvoorbeeld ICOM naam, IP, MAC, softwareversie en disconnectreden."),
    ]
    return pd.DataFrame(source_rows + sheet_rows, columns=["Sectie", "Onderdeel", "Waarde", "Omschrijving"])


def dashboard(writer, kpis, top_ecu, top_fault, timeline):
    kpis.to_excel(excel_writer=writer, sheet_name="Dashboard", index=False, startrow=2)
    top_ecu.to_excel(excel_writer=writer, sheet_name="Dashboard", index=False, startrow=2, startcol=4)
    top_fault.to_excel(excel_writer=writer, sheet_name="Dashboard", index=False, startrow=2, startcol=8)
    timeline.to_excel(excel_writer=writer, sheet_name="Dashboard", index=False, startrow=2, startcol=13)
    ws = writer.book["Dashboard"]
    for addr, val in [("A1", "BMW Diagnose Dashboard"), ("E1", "DTCs per ECU"), ("I1", "Top foutteksten"), ("N1", "DTC tijdlijn")]:
        ws[addr] = val; ws[addr].font = Font(bold=True, size=14, color="1F4E78")
    for col in range(1, 18): ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["B"].width = 32; ws.column_dimensions["J"].width = 45; ws.freeze_panes = "A3"
    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows():
        for cell in row: cell.border = Border(bottom=thin); cell.alignment = Alignment(vertical="top", wrap_text=True)
    if len(top_ecu):
        chart = BarChart(); chart.title = "DTCs per ECU"; chart.y_axis.title = "Aantal"
        chart.add_data(Reference(ws, min_col=6, min_row=3, max_row=3 + len(top_ecu)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=5, min_row=4, max_row=3 + len(top_ecu))); chart.height = 7; chart.width = 12; ws.add_chart(chart, "E16")
    if len(timeline):
        chart = LineChart(); chart.title = "DTC tijdlijn"; chart.y_axis.title = "Aantal"
        chart.add_data(Reference(ws, min_col=15, min_row=3, max_row=3 + len(timeline)), titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=14, min_row=4, max_row=3 + len(timeline))); chart.height = 7; chart.width = 12; ws.add_chart(chart, "N16")


def parse_vehicle_xml_to_excel(trans_xml_path, meta_xml_path, out_excel_path, sort_mode="time", f_ort_width=80, wrap_f_ort_text=True):
    root = parse_xml(trans_xml_path)
    if root is None: raise FileNotFoundError(f"RG_TRANS niet gevonden: {trans_xml_path}")
    vin = (root.findtext("VCI/VIN") or root.findtext("VIN17") or "").strip(); brand = (root.findtext("BrandName") or "").strip()
    rel = related(trans_xml_path, meta_xml_path)
    ecu_df, svk_df, dtc_df, ctx_df, zfs_df = parse_trans(root, vin, brand)
    for col in ["F_UW_ZEIT", "F_UW_KM"]:
        if col not in dtc_df: dtc_df[col] = ""
    dtc_df["F_UW_ZEIT_NUM"] = pd.to_numeric(dtc_df["F_UW_ZEIT"], errors="coerce"); dtc_df["F_UW_KM_NUM"] = pd.to_numeric(dtc_df["F_UW_KM"], errors="coerce")
    sort_cols = ["F_UW_KM_NUM", "F_UW_ZEIT_NUM"] if sort_mode == "km" else ["F_UW_ZEIT_NUM", "F_UW_KM_NUM"]
    dtc_df = dtc_df.sort_values(sort_cols, ascending=[False, False], na_position="last").drop(columns=["F_UW_ZEIT_NUM", "F_UW_KM_NUM"])
    zfs_seconds = dtc_df.groupby(["F_UW_ZEIT", "F_UW_KM"], dropna=False).agg(Speed_Min_kmh=("Speed_Min_kmh", "min"), Speed_Max_kmh=("Speed_Max_kmh", "max"), RPM_Min=("RPM_Min", "min"), RPM_Max=("RPM_Max", "max"), BatteryV_Min=("BatteryV_Min", "min"), BatteryV_Max=("BatteryV_Max", "max"), GPS_Time_UTC=("GPS_Time_UTC", "first"), CustomerTime=("CustomerTime", "first"), DTC_Count=("F_ORT", "count")).reset_index()
    if not zfs_df.empty and "STAT_DM_ZEITSTEMPEL" in zfs_df and "STAT_SYSKONTEXT_KUNDENZEIT" in zfs_df:
        def kz(value):
            try: return (datetime.fromisoformat(str(value).strip()) - datetime(1970, 1, 1)).total_seconds()
            except Exception: return None
        zfs_df["_zfs_num"] = pd.to_numeric(zfs_df["STAT_DM_ZEITSTEMPEL"], errors="coerce"); zfs_df["_kz_unix"] = zfs_df["STAT_SYSKONTEXT_KUNDENZEIT"].apply(kz)
        zfs_df["ZFS_Offset_s"] = (zfs_df["_kz_unix"] - zfs_df["_zfs_num"]).round().astype("Int64")
        calib = zfs_df.dropna(subset=["_zfs_num", "_kz_unix", "ZFS_Offset_s"]).sort_values("_zfs_num")
        if not calib.empty:
            offs = calib["ZFS_Offset_s"].astype(float); calib = calib.assign(Onderbreking_voor_s=offs.diff().fillna(0).where(lambda s: s >= 30, 0).astype(int))
            zfs_df["Onderbreking_voor_s"] = zfs_df["_zfs_num"].map(dict(zip(calib["_zfs_num"], calib["Onderbreking_voor_s"]))).fillna(0).astype(int)
            arr = calib["_zfs_num"].values; off = calib["ZFS_Offset_s"].astype(float).values
            def wall(value):
                try: v = float(value)
                except Exception: return ""
                mask = arr <= v; idx = int(mask.nonzero()[0][-1]) if mask.any() else 0
                return (datetime(1970, 1, 1) + timedelta(seconds=v + off[idx])).strftime("%Y-%m-%d %H:%M:%S")
            dtc_df.insert(dtc_df.columns.get_loc("F_UW_ZEIT") + 1, "Tijdstip_Klant", dtc_df["F_UW_ZEIT"].apply(wall))
        zfs_df = zfs_df.drop(columns=["_zfs_num", "_kz_unix"], errors="ignore")
    prg_ecu, prg_bus, prg_sgbm = parse_prg(rel["prg"], vin)
    fasta_vehicle, fasta_equipment, fasta_ecu, fasta_func = parse_fasta(rel["fstdat"], vin)
    beh_backend, beh_tests, beh_actions, beh_events, beh_device = parse_beh(rel["behdat"], vin)
    status = dtc_df["Context_Status"] if "Context_Status" in dtc_df else pd.Series(dtype="object")
    summary = [("VIN", vin), ("BrandName", brand), ("Script_Version", SCRIPT_VERSION), ("RG_TRANS", Path(trans_xml_path).name), ("RG_META", Path(meta_xml_path).name if meta_xml_path else ""), ("RG_PRG", rel["prg"].name if rel["prg"] else "Niet gevonden"), ("FSTDAT", rel["fstdat"].name if rel["fstdat"] else "Niet gevonden"), ("BEHDAT", rel["behdat"].name if rel["behdat"] else "Niet gevonden"), ("DTC_nodes_in_RG_TRANS", len(root.findall(".//DTC"))), ("DTC_rows_exported", len(dtc_df)), ("DTC_rows_with_context", int((status == "Has typeDTCContext").sum())), ("DTC_rows_empty_context_added", int((status == "Empty DTCContext").sum())), ("DTC_rows_no_context_added", int((status == "No DTCContext").sum())), ("TimeReference", "F_UW_ZEIT is a relative vehicle time counter (ZFS), not epoch/UTC.")] + meta_rows(meta_xml_path)
    summary_df = pd.DataFrame(summary, columns=["Key", "Value"])
    doc_df = documentation_df(rel, trans_xml_path, meta_xml_path)
    top_ecu = dtc_df.groupby("ECUTitle", dropna=False).size().reset_index(name="DTC_Count").sort_values("DTC_Count", ascending=False).head(15) if len(dtc_df) else pd.DataFrame(columns=["ECUTitle", "DTC_Count"])
    top_fault = dtc_df.groupby("F_ORT_TEXT", dropna=False).size().reset_index(name="DTC_Count").sort_values("DTC_Count", ascending=False).head(15) if len(dtc_df) else pd.DataFrame(columns=["F_ORT_TEXT", "DTC_Count"])
    timeline = dtc_df.assign(F_UW_ZEIT_NUM=pd.to_numeric(dtc_df["F_UW_ZEIT"], errors="coerce")).dropna(subset=["F_UW_ZEIT_NUM"]).groupby("F_UW_ZEIT_NUM").size().reset_index(name="DTC_Count").sort_values("F_UW_ZEIT_NUM").tail(40) if len(dtc_df) else pd.DataFrame(columns=["F_UW_ZEIT_NUM", "DTC_Count"])
    kpis = pd.DataFrame([("ECU's", len(ecu_df)), ("DTC-regels", len(dtc_df)), ("Unieke foutlocaties", dtc_df["F_ORT"].nunique() if "F_ORT" in dtc_df else 0), ("DTC's met context", int((status == "Has typeDTCContext").sum())), ("DTC's zonder/lege context toegevoegd", int((status != "Has typeDTCContext").sum())), ("ZFS events", len(zfs_df)), ("FASTA ECU's", len(fasta_ecu)), ("FASTA uitrustingscodes", len(fasta_equipment)), ("BEHDAT acties", len(beh_actions)), ("BEHDAT testmodules NOK", len(beh_tests[beh_tests.get("CollectiveResult", "") != "Ok"]) if len(beh_tests) else 0)], columns=["KPI", "Waarde"])
    sheets = {"Documentatie": doc_df, "Voorblad": summary_df, "ECUs": ecu_df, "DTCs": dtc_df, "DTC_Context": ctx_df, "ZFS_Seconds": zfs_seconds, "ZFS_Resultaten": zfs_df, "SVK_Software": svk_df, "PRG_ECUs": prg_ecu, "PRG_Bus": prg_bus, "PRG_SGBM": prg_sgbm, "FASTA_Vehicle": fasta_vehicle, "FASTA_Equipment": fasta_equipment, "FASTA_ECUs": fasta_ecu, "FASTA_Functions": fasta_func, "BEH_Backend": beh_backend, "BEH_Testmodules": beh_tests, "BEH_Actions": beh_actions, "BEH_Events": beh_events, "BEH_Device": beh_device}
    out = Path(out_excel_path); out.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        dashboard(writer, kpis, top_ecu, top_fault, timeline)
        for name, df in sheets.items():
            (df if len(df) else pd.DataFrame([{"Info": "Geen data gevonden of bronbestand ontbreekt."}])).to_excel(excel_writer=writer, sheet_name=name, index=False)
        for name in writer.book.sheetnames:
            ws = writer.book[name]; style_sheet(ws)
            if False:
                pass
        if "F_ORT_TEXT" in dtc_df:
            ws = writer.book["DTCs"]; col = dtc_df.columns.get_loc("F_ORT_TEXT") + 1; ws.column_dimensions[get_column_letter(col)].width = int(f_ort_width)
            if wrap_f_ort_text:
                for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
                    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical="top")





