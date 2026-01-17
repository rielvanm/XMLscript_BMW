import sys
import time
import traceback
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


OUTPUT_DIR = Path(r"C:\BMW_XML")  # altijd hier opslaan


def strip_namespaces(root: ET.Element) -> None:
    for el in root.iter():
        if "}" in el.tag:
            el.tag = el.tag.split("}", 1)[1]


def safe_float(x: str):
    try:
        return float(x)
    except Exception:
        return None


def update_min_max(cur_min, cur_max, v):
    if v is None:
        return cur_min, cur_max
    if cur_min is None or v < cur_min:
        cur_min = v
    if cur_max is None or v > cur_max:
        cur_max = v
    return cur_min, cur_max


def parse_meta_xml(meta_path: str) -> dict:
    meta_file = Path(meta_path)
    if not meta_file.exists():
        return {}

    tree = ET.parse(meta_file)
    root = tree.getroot()
    strip_namespaces(root)

    return {
        "VIN17": (root.findtext("VIN17") or "").strip(),
        "Baureihe": (root.findtext("BasicFeatures/Baureihe") or "").strip(),
        "Ereihe": (root.findtext("BasicFeatures/Ereihe") or "").strip(),
        "VerkaufsBezeichnung": (root.findtext("BasicFeatures/VerkaufsBezeichnung") or "").strip(),
        "Motor": (root.findtext("BasicFeatures/Motor") or "").strip(),
        "Getriebe": (root.findtext("BasicFeatures/Getriebe") or "").strip(),
        "Modelljahr": (root.findtext("BasicFeatures/Modelljahr") or "").strip(),
        "Marke": (root.findtext("BasicFeatures/Marke") or "").strip(),
        "TypeCode": (root.findtext("BasicFeatures/TypeCode") or "").strip(),
        "StartDate": (root.findtext("StartDate") or "").strip(),
        "EndDate": (root.findtext("EndDate") or "").strip(),
        "WorkState": (root.findtext("WorkState") or "").strip(),
        "VehicleCommunication": (root.findtext("VehicleCommunication") or "").strip(),
        "ComputerName": (root.findtext("ComputerName") or "").strip(),
        "UserName": (root.findtext("UserName") or "").strip(),
        "IstaCaseId": (root.findtext("IstaCaseId") or "").strip(),
        "DateOfFastaRead": (root.findtext("DateOfFastaRead") or "").strip(),
        "DistanceOfFastaRead": (root.findtext("DistanceOfFastaRead") or "").strip(),
        "DealerNumber": (root.findtext("DealerNumber") or "").strip(),
        "distributionPartnerNumber": (root.findtext("distributionPartnerNumber") or "").strip(),
        "outletNumber": (root.findtext("outletNumber") or "").strip(),
    }


def extract_vin(xml_path: str) -> str:
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        strip_namespaces(root)
        return (root.findtext("VCI/VIN") or root.findtext("VIN17") or "").strip()
    except Exception:
        return ""


def pause_exit(msg: str, seconds: int = 6) -> None:
    print(msg)
    print(f"\nVenster sluit automatisch over {seconds} seconden...")
    time.sleep(seconds)


def write_log(out_dir: Path, text: str) -> None:
    try:
        out_dir.mkdir(parents=True, exist_ok=True)
        (out_dir / "BMW_XML_debug.log").write_text(text, encoding="utf-8", errors="replace")
    except Exception:
        pass


def pick_paths_from_argv(argv: list[str]) -> tuple[str, str]:
    meta = ""
    trans = ""
    for a in argv:
        p = Path(a)
        name = p.name.upper()
        if "RG_META" in name and p.suffix.lower() == ".xml":
            meta = str(p)
        if "RG_TRANS" in name and p.suffix.lower() == ".xml":
            trans = str(p)
    return meta, trans


def parse_vehicle_xml_to_excel(
    trans_xml_path: str,
    meta_xml_path: str,
    out_excel_path: str,
    sort_mode: str = "time",  # time/km
    f_ort_width: int = 80,
    wrap_f_ort_text: bool = True,
) -> None:
    trans_file = Path(trans_xml_path)
    if not trans_file.exists():
        raise FileNotFoundError(f"RG_TRANS niet gevonden: {trans_file}")

    tree = ET.parse(trans_file)
    root = tree.getroot()
    strip_namespaces(root)

    vin = (root.findtext("VCI/VIN") or root.findtext("VIN17") or "").strip()
    brand = (root.findtext("BrandName") or "").strip()

    meta = parse_meta_xml(meta_xml_path) if meta_xml_path else {}

    ecu_rows: list[dict] = []
    dtc_rows: list[dict] = []
    dtc_ctx_rows: list[dict] = []

    for ecu in root.findall("ECU/ECU"):
        ec_title = ecu.findtext("ECUTitle", default="")
        variante = ecu.findtext("VARIANTE", default="")
        bus = ecu.findtext("BUS", default="")
        ecu_grobname = ecu.findtext("ECU_GROBNAME", default="")
        ecu_name = ecu.findtext("ECU_NAME", default="")
        ecu_sgbd = ecu.findtext("ECU_SGBD", default="")
        ecu_gruppe = ecu.findtext("ECU_GRUPPE", default="")
        ecu_adr = ecu.findtext("ECU_ADR", default="")
        diag_proto = ecu.findtext("DiagProtocoll", default="")

        comm_ok = ecu.findtext("COMMUNICATION_SUCCESSFULLY", default="")
        ident_ok = ecu.findtext("IDENT_SUCCESSFULLY", default="")
        fs_ok = ecu.findtext("FS_SUCCESSFULLY", default="")
        is_ok = ecu.findtext("IS_SUCCESSFULLY", default="")
        serial_ok = ecu.findtext("SERIAL_SUCCESSFULLY", default="")
        svk_ok = ecu.findtext("SVK_SUCCESSFULLY", default="")

        id_lief_nr = ecu.findtext("ID_LIEF_NR", default="")
        id_lief_text = ecu.findtext("ID_LIEF_TEXT", default="")
        seriennummer = ecu.findtext("SERIENNUMMER", default="")

        prog_datum, prog_km = "", ""
        svk = ecu.find("SVK")
        if svk is not None:
            prog_datum = svk.findtext("PROG_DATUM", default="")
            prog_km = svk.findtext("PROG_KM", default="")

        ecu_rows.append(
            {
                "VIN": vin,
                "BrandName": brand,
                "ECUTitle": ec_title,
                "VARIANTE": variante,
                "ECU_GROBNAME": ecu_grobname,
                "ECU_NAME": ecu_name,
                "ECU_SGBD": ecu_sgbd,
                "ECU_GRUPPE": ecu_gruppe,
                "BUS": bus,
                "ECU_ADR": ecu_adr,
                "DiagProtocoll": diag_proto,
                "ID_LIEF_NR": id_lief_nr,
                "ID_LIEF_TEXT": id_lief_text,
                "SERIENNUMMER": seriennummer,
                "PROG_DATUM": prog_datum,
                "PROG_KM": prog_km,
                "COMMUNICATION_SUCCESSFULLY": comm_ok,
                "IDENT_SUCCESSFULLY": ident_ok,
                "FS_SUCCESSFULLY": fs_ok,
                "IS_SUCCESSFULLY": is_ok,
                "SERIAL_SUCCESSFULLY": serial_ok,
                "SVK_SUCCESSFULLY": svk_ok,
            }
        )

        def add_dtcs(container_tag: str):
            container = ecu.find(container_tag)
            if container is None:
                return

            for dtc in container.findall("DTC"):
                ecu_addr = dtc.findtext("ecuAddress", default="")
                dtc_id = dtc.findtext("dtcId", default="")
                f_ort = dtc.findtext("F_ORT", default="")
                f_ort_text = dtc.findtext("F_ORT_TEXT", default="")
                f_art = dtc.findtext("F_ART", default="")
                f_vorhanden_text = dtc.findtext("F_VORHANDEN_TEXT", default="")
                f_warnung_text = dtc.findtext("F_WARNUNG_TEXT", default="")
                relevance = dtc.findtext("Relevance", default="")
                id_field = dtc.findtext("Id", default="")

                dtc_context = dtc.find("DTCContext")

                # --- CASE 1: Geen context -> wél 1 DTC-regel wegschrijven ---
                if dtc_context is None:
                    dtc_rows.append(
                        {
                            "F_UW_ZEIT": "",
                            "F_UW_KM": "",
                            "F_ORT_TEXT": f_ort_text,
                            "Speed_Min_kmh": None,
                            "Speed_Max_kmh": None,
                            "RPM_Min": None,
                            "RPM_Max": None,
                            "BatteryV_Min": None,
                            "BatteryV_Max": None,
                            "GPS_Time_UTC": None,
                            "CustomerTime": None,
                            "TimeSinceStart_s": None,
                            "Throttle_pct": None,
                            "EngineTemp_C": None,
                            "IntakeTemp_C": None,
                            "Occurrence": None,
                            "F_UW_ANZ": "",
                            "VIN": vin,
                            "BrandName": brand,
                            "ECUTitle": ec_title,
                            "ECU_GROBNAME": ecu_grobname,
                            "ECU_NAME": ecu_name,
                            "VARIANTE": variante,
                            "ECU_ADR": ecu_adr,
                            "Container": container_tag,
                            "ecuAddress": ecu_addr,
                            "dtcId": dtc_id,
                            "Id": id_field,
                            "Relevance": relevance,
                            "F_ORT": f_ort,
                            "F_ART": f_art,
                            "F_VORHANDEN_TEXT": f_vorhanden_text,
                            "F_WARNUNG_TEXT": f_warnung_text,
                        }
                    )
                    # BELANGRIJK: niet 'return' (anders stop je alle volgende DTC's in dit containerblok)
                    continue

                # Meerdere occurrences per DTC
                for occ_idx, type_ctx in enumerate(dtc_context.findall("typeDTCContext"), start=1):
                    ctx_km = type_ctx.findtext("F_UW_KM", default="")
                    ctx_zeit = type_ctx.findtext("F_UW_ZEIT", default="")
                    ctx_anz = type_ctx.findtext("F_UW_ANZ", default="")

                    # reset per occurrence
                    speed_min = speed_max = None
                    rpm_min = rpm_max = None
                    batt_min = batt_max = None

                    gps_time = None
                    customer_time = None

                    tnse_s = None
                    throttle_pct = None
                    tmot_c = None
                    tans_c = None

                    for uw in type_ctx.findall(".//F_UW"):
                        uw_nr = uw.findtext("F_UW_NR", default="")
                        text = uw.findtext("F_UW_TEXT", default="")
                        val_str = uw.findtext("F_UW_WERT", default="")
                        einh = uw.findtext("F_UW_EINH", default="")

                        dtc_ctx_rows.append(
                            {
                                "VIN": vin,
                                "BrandName": brand,
                                "ECUTitle": ec_title,
                                "Container": container_tag,
                                "ecuAddress": ecu_addr,
                                "dtcId": dtc_id,
                                "Id": id_field,
                                "Occurrence": occ_idx,
                                "F_UW_KM": ctx_km,
                                "F_UW_ZEIT": ctx_zeit,
                                "F_UW_ANZ": ctx_anz,
                                "F_UW_NR": uw_nr,
                                "F_UW_TEXT": text,
                                "F_UW_WERT": val_str,
                                "F_UW_EINH": einh,
                            }
                        )

                        t = (text or "").lower()
                        einh_l = (einh or "").lower()

                        if gps_time is None and ("gps" in t and "zeit" in t):
                            gps_time = val_str
                        if customer_time is None and ("kundenzeit" in t or "customer" in t):
                            customer_time = val_str

                        v = safe_float(val_str)
                        if v is None:
                            continue

                        if (("geschwindigkeit" in t or "speed" in t) and "km/h" in einh_l):
                            speed_min, speed_max = update_min_max(speed_min, speed_max, v)

                        if ("motordrehzahl" in t) or ("nmot" in t) or ("rpm" in t and "min" in einh_l):
                            rpm_min, rpm_max = update_min_max(rpm_min, rpm_max, v)

                        if ("batteriespannung" in t) or ("bordnetz" in t) or ("ub_w" in t) or (
                            "voltage" in t and "v" in einh_l
                        ):
                            batt_min, batt_max = update_min_max(batt_min, batt_max, v)

                        if tnse_s is None and ("zeit nach motorstart" in t or "tnse" in t):
                            tnse_s = v

                        if throttle_pct is None and ("fahrwertgeber" in t or "fwg" in t or "throttle" in t):
                            throttle_pct = v

                        if tmot_c is None and ("motortemperatur" in t or "tmot" in t) and ("c" in einh_l or "°" in einh_l):
                            tmot_c = v
                        if tans_c is None and ("ansauglufttemperatur" in t or "tans" in t or "intake" in t) and ("c" in einh_l or "°" in einh_l):
                            tans_c = v

                    dtc_rows.append(
                        {
                            "F_UW_ZEIT": ctx_zeit,
                            "F_UW_KM": ctx_km,
                            "F_ORT_TEXT": f_ort_text,
                            "F_UW_ANZ": ctx_anz,
                            "Speed_Min_kmh": speed_min,
                            "Speed_Max_kmh": speed_max,
                            "RPM_Min": rpm_min,
                            "RPM_Max": rpm_max,
                            "BatteryV_Min": batt_min,
                            "BatteryV_Max": batt_max,
                            "GPS_Time_UTC": gps_time,
                            "CustomerTime": customer_time,
                            "TimeSinceStart_s": tnse_s,
                            "Throttle_pct": throttle_pct,
                            "EngineTemp_C": tmot_c,
                            "IntakeTemp_C": tans_c,
                            "ECUTitle": ec_title,
                            "ECU_NAME": ecu_name,
                            "VARIANTE": variante,
                            "ECU_ADR": ecu_adr,
                            "Container": container_tag,
                            "ecuAddress": ecu_addr,
                            "dtcId": dtc_id,
                            "Id": id_field,
                            "F_ORT": f_ort,
                            "F_ART": f_art,
                            "F_VORHANDEN_TEXT": f_vorhanden_text,
                            "F_WARNUNG_TEXT": f_warnung_text,
                            "Relevance": relevance,
                            "Occurrence": occ_idx,
                            "VIN": vin,
                            "BrandName": brand,
                            "ECU_GROBNAME": ecu_grobname,
                        }
                    )

        add_dtcs("FEHLER")
        add_dtcs("INFO")

    ecu_df = pd.DataFrame(ecu_rows)
    dtc_df = pd.DataFrame(dtc_rows)
    ctx_df = pd.DataFrame(dtc_ctx_rows)

    dtc_df["F_UW_ZEIT_NUM"] = pd.to_numeric(dtc_df["F_UW_ZEIT"], errors="coerce")
    dtc_df["F_UW_KM_NUM"] = pd.to_numeric(dtc_df["F_UW_KM"], errors="coerce")

    if sort_mode == "km":
        dtc_df = dtc_df.sort_values(by=["F_UW_KM_NUM", "F_UW_ZEIT_NUM"], ascending=[False, False], na_position="last")
    else:
        dtc_df = dtc_df.sort_values(by=["F_UW_ZEIT_NUM", "F_UW_KM_NUM"], ascending=[False, False], na_position="last")

    dtc_df = dtc_df.drop(columns=["F_UW_ZEIT_NUM", "F_UW_KM_NUM"])

    zfs_df = (
        dtc_df.groupby(["F_UW_ZEIT", "F_UW_KM"], dropna=False)
        .agg(
            Speed_Min_kmh=("Speed_Min_kmh", "min"),
            Speed_Max_kmh=("Speed_Max_kmh", "max"),
            RPM_Min=("RPM_Min", "min"),
            RPM_Max=("RPM_Max", "max"),
            BatteryV_Min=("BatteryV_Min", "min"),
            BatteryV_Max=("BatteryV_Max", "max"),
            GPS_Time_UTC=("GPS_Time_UTC", "first"),
            CustomerTime=("CustomerTime", "first"),
            DTC_Count=("dtcId", "count"),
        )
        .reset_index()
    )

    summary = {
        "VIN": vin,
        "BrandName": brand,
        "TimeReference": "F_UW_ZEIT is a relative vehicle time counter (ZFS), not epoch/UTC.",
    }
    for k, v in (meta or {}).items():
        summary[f"Meta_{k}"] = v

    summary_df = pd.DataFrame(list(summary.items()), columns=["Key", "Value"])

    out_path = Path(out_excel_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Voorblad", index=False)
        ecu_df.to_excel(writer, sheet_name="ECUs", index=False)
        dtc_df.to_excel(writer, sheet_name="DTCs", index=False)
        ctx_df.to_excel(writer, sheet_name="DTC_Context", index=False)
        zfs_df.to_excel(writer, sheet_name="ZFS_Seconds", index=False)

        ws = writer.book["DTCs"]
        if "F_ORT_TEXT" in dtc_df.columns:
            col_idx = dtc_df.columns.get_loc("F_ORT_TEXT") + 1
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = int(f_ort_width)

            if wrap_f_ort_text:
                alignment = Alignment(wrap_text=True, vertical="top")
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = alignment


if __name__ == "__main__":
    args = sys.argv[1:]
    meta_xml, trans_xml = pick_paths_from_argv(args)

    if not meta_xml or not trans_xml:
        pause_exit(
            "Sleep ZOWEL RG_META*.xml ALS RG_TRANS*.xml op deze EXE.\n"
            "Of start vanuit cmd met:\n"
            "  BMW_XML.exe <pad_naar_RG_META.xml> <pad_naar_RG_TRANS.xml>",
            seconds=10,
        )
        raise SystemExit(2)

    try:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        vin = extract_vin(trans_xml) or extract_vin(meta_xml) or "UNKNOWNVIN"
        out_excel = OUTPUT_DIR / f"BMW_XML_{vin}.xlsx"

        print("Meta:", meta_xml)
        print("Trans:", trans_xml)
        print("Output:", out_excel)

        parse_vehicle_xml_to_excel(
            trans_xml_path=trans_xml,
            meta_xml_path=meta_xml,
            out_excel_path=str(out_excel),
        )

        if out_excel.exists():
            pause_exit(f"Klaar! Excel gemaakt:\n{out_excel}", seconds=10)
        else:
            pause_exit(f"Fout: schrijven leek te slagen, maar bestand bestaat niet:\n{out_excel}", seconds=10)

    except Exception as e:
        tb = traceback.format_exc()
        write_log(OUTPUT_DIR, tb)
        pause_exit(f"Fout: {e}\n\nDetails staan in:\n{OUTPUT_DIR}\\BMW_XML_debug.log", seconds=15)
        raise
